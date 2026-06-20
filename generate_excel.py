"""
Generate Excel review spreadsheet for the Notion onboarding template.
Tabs:
  1. Lab Culture & Conduct
  2. Foundations & Setup
  3. Technical Knowhow
  4. Funding & Fellowships
  5. Projects
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ---------- Styling ----------
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
RATING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
REMARKS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# Category color palette — each category gets a distinct soft color
CATEGORY_COLORS = [
    "DCE6F1",  # light blue
    "E2EFDA",  # light green
    "FCE4D6",  # light orange
    "D9E2F3",  # periwinkle
    "FFF2CC",  # light yellow
    "F2DCDB",  # light pink
    "DDEBF7",  # sky blue
    "E4DFEC",  # light purple
    "D6E4F0",  # steel blue
    "C6EFCE",  # mint green
    "FDE9D9",  # peach
    "DAEEF3",  # light teal
    "F2F2F2",  # light gray
    "E6E0EC",  # lavender
    "FFCCCC",  # salmon
]


def get_cat_fill(cat_name, color_map):
    """Get or assign a color for a category."""
    if cat_name not in color_map:
        idx = len(color_map) % len(CATEGORY_COLORS)
        color_map[cat_name] = PatternFill(
            start_color=CATEGORY_COLORS[idx],
            end_color=CATEGORY_COLORS[idx],
            fill_type="solid",
        )
    return color_map[cat_name]


def setup_sheet(ws, title):
    """Set up a sheet with standard headers, tab rating row, and column widths."""
    ws.title = title

    # --- Column widths ---
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    for col_letter in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 40

    LOCKED_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    # --- Row 1: Tab-level rating ---
    TAB_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    TAB_FONT = Font(bold=True, size=12, color="FFFFFF")
    LOCKED_FILL_GRAY = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

    cell_a = ws.cell(row=1, column=1, value=f'Tab: "{title}"')
    cell_a.font = TAB_FONT
    cell_a.fill = TAB_FILL
    cell_a.alignment = Alignment(vertical="center")
    cell_a.border = THIN_BORDER

    cell_b = ws.cell(row=1, column=2, value="Rate this tab overall →")
    cell_b.font = Font(bold=True, size=10, color="FFFFFF")
    cell_b.fill = TAB_FILL
    cell_b.alignment = Alignment(horizontal="right", vertical="center")
    cell_b.border = THIN_BORDER

    for col_idx in range(3, 5):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = TAB_FILL
        cell.border = THIN_BORDER

    for col_idx in range(5, 10):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = RATING_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    avg_cell = ws.cell(row=1, column=10)
    avg_cell.value = '=IF(COUNT(E1:I1)>0,AVERAGE(E1:I1),"")'
    avg_cell.border = THIN_BORDER
    avg_cell.alignment = Alignment(horizontal="center")
    avg_cell.number_format = "0.0"

    cell_k = ws.cell(row=1, column=11, value="Comments on tab name / scope")
    cell_k.fill = REMARKS_FILL
    cell_k.border = THIN_BORDER
    cell_k.alignment = WRAP

    ws.row_dimensions[1].height = 30

    # --- Row 2: Column headers ---
    headers = [
        "Category", "Task / Item", "Type", "Status",
        "Person 1\nRating (1-5)", "Person 2\nRating (1-5)", "Person 3\nRating (1-5)",
        "Person 4\nRating (1-5)", "Person 5\nRating (1-5)", "Average\nRating",
        "Remarks"
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 35
    ws.freeze_panes = "A3"

    return LOCKED_FILL_GRAY


def add_tasks(ws, data, start_row=3):
    """Add category/task rows with color-coded categories.
    Inserts a category-level rating row before each new category group.
    Non-writable cells on task rows are grayed out.
    data = [(category, task_name, type_label, status), ...]
    """
    LOCKED_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    LOCKED_FONT = Font(color="666666")
    CAT_HEADER_FONT = Font(bold=True, size=11, color="000000")
    color_map = {}
    row = start_row
    current_cat = None

    for category, task, type_label, status in data:
        cat_fill = get_cat_fill(category, color_map)

        # Insert category rating row when category changes
        if category != current_cat:
            current_cat = category
            # Category header row
            cell_a = ws.cell(row=row, column=1, value=category)
            cell_a.font = CAT_HEADER_FONT
            cell_a.fill = cat_fill
            cell_a.border = THIN_BORDER
            cell_a.alignment = WRAP

            cell_b = ws.cell(row=row, column=2, value="⬅ Rate this category →")
            cell_b.font = Font(bold=True, italic=True, size=10, color="555555")
            cell_b.fill = cat_fill
            cell_b.border = THIN_BORDER
            cell_b.alignment = Alignment(horizontal="right", vertical="top")

            for col_idx in range(3, 5):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = LOCKED_FILL
                cell.border = THIN_BORDER

            for col_idx in range(5, 10):
                cell = ws.cell(row=row, column=col_idx)
                cell.fill = RATING_FILL
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            avg_cell = ws.cell(row=row, column=10)
            avg_cell.value = f'=IF(COUNT(E{row}:I{row})>0,AVERAGE(E{row}:I{row}),"")'
            avg_cell.border = THIN_BORDER
            avg_cell.alignment = Alignment(horizontal="center")
            avg_cell.number_format = "0.0"

            remarks_cell = ws.cell(row=row, column=11)
            remarks_cell.fill = REMARKS_FILL
            remarks_cell.border = THIN_BORDER
            remarks_cell.alignment = WRAP

            row += 1

        # Task row
        cell_a = ws.cell(row=row, column=1, value="")
        cell_a.fill = LOCKED_FILL
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=row, column=2, value=task)
        cell_b.border = THIN_BORDER
        cell_b.alignment = WRAP

        cell_c = ws.cell(row=row, column=3, value=type_label)
        cell_c.fill = LOCKED_FILL
        cell_c.font = LOCKED_FONT
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center", vertical="top")

        cell_d = ws.cell(row=row, column=4, value=status)
        cell_d.fill = LOCKED_FILL
        cell_d.font = LOCKED_FONT
        cell_d.border = THIN_BORDER
        cell_d.alignment = Alignment(horizontal="center", vertical="top")

        # Rating columns (E-I)
        for col_idx in range(5, 10):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = RATING_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Average formula (J)
        avg_cell = ws.cell(row=row, column=10)
        avg_cell.value = f'=IF(COUNT(E{row}:I{row})>0,AVERAGE(E{row}:I{row}),"")'
        avg_cell.border = THIN_BORDER
        avg_cell.alignment = Alignment(horizontal="center")
        avg_cell.number_format = "0.0"

        # Remarks (K)
        remarks_cell = ws.cell(row=row, column=11)
        remarks_cell.fill = REMARKS_FILL
        remarks_cell.border = THIN_BORDER
        remarks_cell.alignment = WRAP

        row += 1
    return row


# ================================================================
# TAB 1: Lab Culture & Conduct
# ================================================================
ws1 = wb.active
setup_sheet(ws1, "Lab Culture & Conduct")

handbook_data = [
    ("Lab Orientation", "Lab Introduction", "Theory", "Existing"),
    ("Lab Orientation", "Student Role & Expectations", "Theory", "Existing"),
    ("Lab Orientation", "Postdoc Role", "Theory", "Existing"),
    ("Lab Orientation", "PI Role", "Theory", "Existing"),
    ("Workplace Conduct", "Conduct in Meetings", "Theory", "Existing"),
    ("Workplace Conduct", "Office Interaction Guidelines", "Theory", "Existing"),
    ("Workplace Conduct", "Socializing Guidelines", "Theory", "Existing"),
    ("Workplace Conduct", "Inappropriate Behavior Policy", "Theory", "Existing"),
    ("Work-Life Balance", "Setting Clear Expectations", "Theory", "Existing"),
    ("Work-Life Balance", "Work-Life Balance & Wellbeing", "Theory", "Existing"),
    ("Work-Life Balance", "Equality, Diversity & Inclusion", "Theory", "Existing"),
    ("Professional Development", "Create Individual Development Plan (IDP)", "Theory", "Existing"),
    ("Professional Development", "Prepare for Annual Evaluations", "Theory", "Existing"),
    ("Research Practices", "Reproducible Research", "Theory", "Existing"),
    ("Research Practices", "Research Conduct & Ethics", "Theory", "Existing"),
    ("Collaboration & Conferences", "Collaboration Guidelines", "Theory", "Existing"),
    ("Collaboration & Conferences", "Poster Printing Procedures", "Theory", "Existing"),
    ("Collaboration & Conferences", "Conference Expectations & Logistics", "Theory", "Existing"),
]
add_tasks(ws1, handbook_data)


# ================================================================
# TAB 2: Foundations & Setup
# ================================================================
ws2 = wb.create_sheet()
setup_sheet(ws2, "Foundations & Setup")

tech_data = [
    ("Research Foundations", "Understand What fMRI Measures", "Theory", "Existing"),
    ("Research Foundations", "Learn Basic Brain Anatomy for Neuroimaging", "Theory", "Existing"),
    ("Research Foundations", "Learn What ICA Is and Why It Matters", "Theory", "Existing"),
    ("Research Foundations", "Understand Group ICA for Multi-Subject Studies", "Theory", "Existing"),
    ("Research Foundations", "Learn Brain Networks", "Theory", "Existing"),
    ("Research Foundations", "Read Key Lab Papers", "Theory", "Existing"),
    ("Research Foundations", "Learn Common Analysis Patterns in the Lab", "Theory", "Existing"),
    ("Research Foundations", "Interpret Functional Connectivity Matrices", "Theory", "Existing"),
    ("Research Foundations", "Understand Why fMRI Preprocessing Matters", "Theory", "Existing"),
    ("Research Foundations", "Learn Quality Control for fMRI Data", "Theory", "Existing"),
    ("Statistical Analysis", "Understand the General Linear Model (GLM)", "Theory", "Existing"),
    ("Statistical Analysis", "Learn Group Statistics (t-tests, ANOVA)", "Theory", "Existing"),
    ("Statistical Analysis", "Understand Statistics for Neuroimaging", "Theory", "Existing"),
    ("Neuroimaging Software", "Survey the Neuroimaging Software Landscape", "Theory", "Existing"),
    ("Writing & Publishing", "Learn Scientific Writing Basics", "Theory", "Existing"),
    ("Writing & Publishing", "Practice Reproducibility and Open Science", "Theory", "Existing"),
    ("Computing Environment", "Cluster Access & Connectivity", "Hands-On", "Existing"),
    ("Computing Environment", "Storage & File Management", "Hands-On", "Existing"),
    ("Computing Environment", "SLURM Job Submission", "Hands-On", "Existing"),
    ("Analysis Methods", "Install & Launch GIFT Toolbox", "Hands-On", "Existing"),
    ("Analysis Methods", "Run Your First Group ICA Analysis", "Hands-On", "Existing"),
    ("Analysis Methods", "Visualize & Sort ICA Components", "Hands-On", "Existing"),
    ("Analysis Methods", "Debug and Troubleshoot ICA Analyses", "Hands-On", "Existing"),
    ("Analysis Methods", "Run Single-Subject Neuromark ICA on Cluster", "Hands-On", "Existing"),
    ("Analysis Methods", "Run Multi-Subject ICA with Array Jobs", "Hands-On", "Existing"),
]
add_tasks(ws2, tech_data)


# ================================================================
# TAB 3: Technical Knowhow
# ================================================================
ws3 = wb.create_sheet()
setup_sheet(ws3, "Tools & Workflows")

knowhow_data = [
    ("Git & GitHub", "", "Hands-On", "Proposed"),
    ("Git & GitHub", "", "Hands-On", "Proposed"),
    ("Git & GitHub", "", "Hands-On", "Proposed"),
    ("VS Code", "", "Hands-On", "Proposed"),
    ("VS Code", "", "Hands-On", "Proposed"),
    ("VS Code", "", "Hands-On", "Proposed"),
    ("MATLAB", "", "Hands-On", "Proposed"),
    ("MATLAB", "", "Hands-On", "Proposed"),
    ("MATLAB", "", "Hands-On", "Proposed"),
]
add_tasks(ws3, knowhow_data)


# ================================================================
# TAB 4: Funding & Fellowships
# (Empty placeholder rows — professor fills in real items)
# ================================================================
ws4 = wb.create_sheet()
setup_sheet(ws4, "Funding & Fellowships")

funding_data = [
    ("Grants", "", "Theory", ""),
    ("Grants", "", "Theory", ""),
    ("Grants", "", "Theory", ""),
    ("Fellowships", "", "Theory", ""),
    ("Fellowships", "", "Theory", ""),
    ("Fellowships", "", "Theory", ""),
    ("Travel Awards", "", "Theory", ""),
    ("Travel Awards", "", "Theory", ""),
    ("Travel Awards", "", "Theory", ""),
]
add_tasks(ws4, funding_data)


# ================================================================
# TAB 5: Projects (uses same setup_sheet + add_tasks pattern)
# ================================================================
ws5 = wb.create_sheet()
setup_sheet(ws5, "Projects")

projects_data = [
    ("Capstone Project", "Capstone: Schizophrenia vs Controls Analysis (OpenNeuro ds000030)", "Hands-On", "Existing"),
]
add_tasks(ws5, projects_data)


# Save
output_path = r"C:\Notion\Onboarding_Template_Review.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
