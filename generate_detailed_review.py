"""
Generate a detailed review spreadsheet for Notion onboarding content.
Each task gets a row; each reviewer gets a column for comments.
5 reviewers: Armin, Spencer, Kyle, Prerana, Nigar.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

REVIEWERS = ["Armin", "Spencer", "Kyle", "Prerana", "Nigar"]

# ---------- Styling ----------
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CAT_FONT = Font(bold=True, size=11, color="000000")
COMMENT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LOCKED_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

CATEGORY_COLORS = [
    "DCE6F1", "E2EFDA", "FCE4D6", "D9E2F3", "FFF2CC",
    "F2DCDB", "DDEBF7", "E4DFEC", "D6E4F0", "C6EFCE",
    "FDE9D9", "DAEEF3", "F2F2F2", "E6E0EC", "FFCCCC",
]


def get_cat_fill(cat_name, color_map):
    if cat_name not in color_map:
        idx = len(color_map) % len(CATEGORY_COLORS)
        color_map[cat_name] = PatternFill(
            start_color=CATEGORY_COLORS[idx],
            end_color=CATEGORY_COLORS[idx],
            fill_type="solid",
        )
    return color_map[cat_name]


def setup_sheet(ws, title):
    ws.title = title

    # Column widths: A=Category, B=Task, C=Type, D..H=Reviewer comments
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    for i, _ in enumerate(REVIEWERS):
        col_letter = get_column_letter(4 + i)  # D, E, F, G, H
        ws.column_dimensions[col_letter].width = 40

    # Row 1: Headers
    headers = ["Category", "Task / Item", "Type"] + [
        f"{name}\nComments" for name in REVIEWERS
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"


def add_tasks(ws, data, start_row=2):
    """Add rows. data = [(category, task_name, type_label), ...]"""
    color_map = {}
    row = start_row
    current_cat = None

    for category, task, type_label in data:
        cat_fill = get_cat_fill(category, color_map)

        # Category header row
        if category != current_cat:
            current_cat = category
            cell_a = ws.cell(row=row, column=1, value=category)
            cell_a.font = CAT_FONT
            cell_a.fill = cat_fill
            cell_a.border = THIN_BORDER
            cell_a.alignment = WRAP

            cell_b = ws.cell(row=row, column=2)
            cell_b.fill = cat_fill
            cell_b.border = THIN_BORDER

            cell_c = ws.cell(row=row, column=3)
            cell_c.fill = cat_fill
            cell_c.border = THIN_BORDER

            # Gray out reviewer columns on category rows
            for i in range(len(REVIEWERS)):
                cell = ws.cell(row=row, column=4 + i)
                cell.fill = LOCKED_FILL
                cell.border = THIN_BORDER
            row += 1

        # Task row
        cell_a = ws.cell(row=row, column=1)
        cell_a.fill = LOCKED_FILL
        cell_a.border = THIN_BORDER

        cell_b = ws.cell(row=row, column=2, value=task)
        cell_b.border = THIN_BORDER
        cell_b.alignment = WRAP

        cell_c = ws.cell(row=row, column=3, value=type_label)
        cell_c.fill = LOCKED_FILL
        cell_c.border = THIN_BORDER
        cell_c.alignment = Alignment(horizontal="center", vertical="top")

        # Comment columns — writable, green
        for i in range(len(REVIEWERS)):
            cell = ws.cell(row=row, column=4 + i)
            cell.fill = COMMENT_FILL
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        ws.row_dimensions[row].height = 60
        row += 1

    return row


# ================================================================
# TAB 0: Welcome (landing page — single page, not a database)
# ================================================================
ws0 = wb.active
setup_sheet(ws0, "Welcome")

welcome_data = [
    ("Welcome Page", "Welcome To Image-Analysis Lab at TReNDS Center", "Page"),
    ("Welcome Page", "Lab Introduction & Purpose", "Section"),
    ("Welcome Page", "Lab Leadership", "Section"),
]
add_tasks(ws0, welcome_data)


# ================================================================
# TAB 1: Lab Culture & Conduct
# ================================================================
ws1 = wb.create_sheet()
setup_sheet(ws1, "Lab Culture & Conduct")

handbook_data = [
    ("Lab Orientation", "Lab Introduction", "Theory"),
    ("Lab Orientation", "Student Role & Expectations", "Theory"),
    ("Lab Orientation", "Postdoc Role", "Theory"),
    ("Lab Orientation", "PI Role", "Theory"),
    ("Workplace Conduct", "Conduct in Meetings", "Theory"),
    ("Workplace Conduct", "Office Interaction Guidelines", "Theory"),
    ("Workplace Conduct", "Socializing Guidelines", "Theory"),
    ("Workplace Conduct", "Inappropriate Behavior Policy", "Theory"),
    ("Work-Life Balance", "Setting Clear Expectations", "Theory"),
    ("Work-Life Balance", "Work-Life Balance & Wellbeing", "Theory"),
    ("Work-Life Balance", "Equality, Diversity & Inclusion", "Theory"),
    ("Professional Development", "Create Individual Development Plan (IDP)", "Theory"),
    ("Professional Development", "Prepare for Annual Evaluations", "Theory"),
    ("Research Practices", "Reproducible Research", "Theory"),
    ("Research Practices", "Research Conduct & Ethics", "Theory"),
    ("Collaboration & Conferences", "Collaboration Guidelines", "Theory"),
    ("Collaboration & Conferences", "Poster Printing Procedures", "Theory"),
    ("Collaboration & Conferences", "Conference Expectations & Logistics", "Theory"),
]
add_tasks(ws1, handbook_data)


# ================================================================
# TAB 2: Foundations & Setup
# ================================================================
ws2 = wb.create_sheet()
setup_sheet(ws2, "Foundations & Setup")

tech_data = [
    ("Getting Started", "Read Technical Onboarding Introduction", "Theory"),
    ("Getting Started", "Set Up and Review TReNDS Computing Cluster", "Hands-On"),
    ("Getting Started", "Learn Independent Component Analysis (ICA)", "Theory"),
    ("Getting Started", "Identify and Label Brain Networks", "Theory"),
    ("Getting Started", "Learn Data Visualization Tools", "Hands-On"),
    ("Getting Started", "Explore TReNDS Databases and Resources", "Hands-On"),
    ("Research Foundations", "Understand What fMRI Measures", "Theory"),
    ("Research Foundations", "Learn Basic Brain Anatomy for Neuroimaging", "Theory"),
    ("Research Foundations", "Learn What ICA Is and Why It Matters", "Theory"),
    ("Research Foundations", "Understand Group ICA for Multi-Subject Studies", "Theory"),
    ("Research Foundations", "Learn Brain Networks", "Theory"),
    ("Research Foundations", "Read Key Lab Papers", "Theory"),
    ("Research Foundations", "Learn Common Analysis Patterns in the Lab", "Theory"),
    ("Research Foundations", "Interpret Functional Connectivity Matrices", "Theory"),
    ("Research Foundations", "Understand Why fMRI Preprocessing Matters", "Theory"),
    ("Research Foundations", "Learn Quality Control for fMRI Data", "Theory"),
    ("Statistical Analysis", "Understand the General Linear Model (GLM)", "Theory"),
    ("Statistical Analysis", "Learn Group Statistics (t-tests, ANOVA)", "Theory"),
    ("Statistical Analysis", "Understand Statistics for Neuroimaging", "Theory"),
    ("Neuroimaging Software", "Survey the Neuroimaging Software Landscape", "Theory"),
    ("Writing & Publishing", "Learn Scientific Writing Basics", "Theory"),
    ("Writing & Publishing", "Practice Reproducibility and Open Science", "Theory"),
    ("Computing Environment", "Cluster Access & Connectivity", "Hands-On"),
    ("Computing Environment", "Storage & File Management", "Hands-On"),
    ("Computing Environment", "SLURM Job Submission", "Hands-On"),
    ("Analysis Methods", "Install & Launch GIFT Toolbox", "Hands-On"),
    ("Analysis Methods", "Run Your First Group ICA Analysis", "Hands-On"),
    ("Analysis Methods", "Visualize & Sort ICA Components", "Hands-On"),
    ("Analysis Methods", "Debug and Troubleshoot ICA Analyses", "Hands-On"),
    ("Analysis Methods", "Run Single-Subject Neuromark ICA on Cluster", "Hands-On"),
    ("Analysis Methods", "Run Multi-Subject ICA with Array Jobs", "Hands-On"),
]
add_tasks(ws2, tech_data)


# ================================================================
# TAB 3: Tools & Workflows
# ================================================================
ws3 = wb.create_sheet()
setup_sheet(ws3, "Tools & Workflows")

tools_data = [
    ("Git & GitHub", "Git & GitHub Basics", "Hands-On"),
    ("VS Code", "VS Code for Remote Development", "Hands-On"),
    ("MATLAB", "MATLAB Fundamentals", "Hands-On"),
]
add_tasks(ws3, tools_data)


# ================================================================
# TAB 4: Funding & Fellowships
# ================================================================
ws4 = wb.create_sheet()
setup_sheet(ws4, "Funding & Fellowships")

funding_data = [
    ("Grants", "Grants Overview", "Theory"),
    ("Fellowships", "Fellowships Overview", "Theory"),
    ("Travel Awards", "Travel Awards Overview", "Theory"),
]
add_tasks(ws4, funding_data)


# ================================================================
# TAB 5: Projects
# ================================================================
ws5 = wb.create_sheet()
setup_sheet(ws5, "Projects")

projects_data = [
    ("Capstone Project", "Capstone: Schizophrenia vs Controls Analysis (OpenNeuro ds000030)", "Hands-On"),
]
add_tasks(ws5, projects_data)


# Save
output_path = r"C:\Notion\Onboarding_Detailed_Review.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
